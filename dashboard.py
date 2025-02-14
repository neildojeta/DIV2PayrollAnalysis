import xlwings as xw
import openpyxl
from xlwings.utils import rgb_to_int
import win32com.client
from openpyxl.drawing.image import Image
from PIL import Image as PILImage, ImageDraw, ImageFont
import os
import logging
import time
import sys
import week_tables as wt

# Set up logging
log_folder = "Logs"
os.makedirs(log_folder, exist_ok=True)
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(log_folder, 'Comparison.log')),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger()

# Define the comparison files and corresponding sheets
def main(file_previous, file_latest): 
    logger.info(f"{file_previous} + {file_latest}")
    comparison_files = [
        ('ComparedResults/DIV2_Tables.xlsx', 'Dashboard')
    ]

    # Open the Dashboard workbook
    dashboard_file = 'ComparedResults/Dashboard.xlsm'
    app = xw.App(visible=True)

    try:
        # Open the existing workbook (Dashboard)
        wb_dashboard = app.books.open(dashboard_file)

        # Loop through each comparison file and corresponding sheet
        for comparison_file, sheet_name in comparison_files:

            # Load the comparison workbook and sheet
            wb_comparison = openpyxl.load_workbook(comparison_file)
            sheet_TotalInvoicePayment = wb_comparison['TotalInvoicePayment']
            sheet_HTotalRevComparison = wb_comparison['HTotalRevComparison']
            sheet_LiftLeaseComparison = wb_comparison['LiftLeaseComparison']
            sheet_ViolationComparison = wb_comparison['ViolationComparison']
            # sheet_OperatorChanges = wb_comparison['OperatorChanges']
            # paste_picture(comparison_files, dashboard_file)

            # Retrieve Payment values from the 'TotalInvoicePayment' sheet in the comparison file
            full_prev_amount = f"{float(sheet_TotalInvoicePayment['A2'].value):,.2f}"
            full_lat_amount = f"{float(sheet_TotalInvoicePayment['B2'].value):,.2f}"

            # Retrieve values from the 'TotalInvoicePayment' sheet in the comparison file
            full_amount_diff = f"{abs(float(sheet_TotalInvoicePayment['C2'].value)):,.2f}"
            full_amount_diff_status = sheet_TotalInvoicePayment['D2'].value

            # Retrieve values from the 'HTotalRevComparison' sheet in the comparison file
            prev_HTotalRev = f"{sum(cell.value for row in sheet_HTotalRevComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
            lat_HTotalRev = f"{sum(cell.value for row in sheet_HTotalRevComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
            diff_HTotalRev = f"{sum(cell.value for row in sheet_HTotalRevComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

            # Retrieve values from the 'LiftLeaseComparison' sheet in the comparison file
            prev_LeaseComp = f"{sum(cell.value for row in sheet_LiftLeaseComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
            lat_LeaseComp = f"{sum(cell.value for row in sheet_LiftLeaseComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
            diff_LeaseComp = f"{sum(cell.value for row in sheet_LiftLeaseComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

            # Retrieve values from the 'ViolationComparison' sheet in the comparison file
            prev_ViolationComp = f"{sum(cell.value for row in sheet_ViolationComparison.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):,.2f}"
            lat_ViolationComp = f"{sum(cell.value for row in sheet_ViolationComparison.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):,.2f}"
            diff_ViolationComp = f"{sum(cell.value for row in sheet_ViolationComparison.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):,.2f}"

            # Retrieve values from the 'OperatorChanges' sheet in the comparison file
            # prev_OperatorChanges = f"{sum(cell.value for row in sheet_OperatorChanges.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None) or 0:.0f}"
            # lat_OperatorChanges = f"{sum(cell.value for row in sheet_OperatorChanges.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None) or 0:.0f}"
            # diff_OperatorChanges = f"{sum(cell.value for row in sheet_OperatorChanges.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None) or 0:.0f}"

            # Get the corresponding sheet in the dashboard
            sheet_dashboard = wb_dashboard.sheets[sheet_name]

            # Access the file name values shape via the API and set the value
            txt_prev_file = sheet_dashboard.shapes['txtPrevFile'].api
            txt_prev_file.TextFrame2.TextRange.Text = f"{file_previous}"

            txt_lat_file = sheet_dashboard.shapes['txtLatFile'].api
            txt_lat_file.TextFrame2.TextRange.Text = f"{file_latest}"

            # Access the total diff value shape via the API and set the value
            txt_full_amount_diff = sheet_dashboard.shapes['TextBox 88'].api
            txt_full_amount_diff.TextFrame2.TextRange.Text = f"$ {full_amount_diff}"

            txt_full_prev_amount = sheet_dashboard.shapes['txtPrevPAyment'].api
            txt_full_prev_amount.TextFrame2.TextRange.Text = f"$ {full_prev_amount}"

            txt_full_lat_amount = sheet_dashboard.shapes['txtLatPayment'].api
            txt_full_lat_amount.TextFrame2.TextRange.Text = f"$ {full_lat_amount}"

            # Access the total diff status shape via the API and set the value
            txt_full_amount_diff_status = sheet_dashboard.shapes['TextBox 90'].api
            txt_full_amount_diff_status.TextFrame2.TextRange.Text = f"{full_amount_diff_status}"

            # Access the HTotalRevComparison shape via the API and set the value
            txt_HTotalRev = sheet_dashboard.shapes['txtDHTotalRevDiff'].api
            txt_HTotalRev.TextFrame2.TextRange.Text = f"${prev_HTotalRev} to ${lat_HTotalRev}"
            txt_HTotalRev_diff = sheet_dashboard.shapes['txtHTotalRevDiff'].api
            txt_HTotalRev_diff.TextFrame2.TextRange.Text = f"${diff_HTotalRev}"

            # Access the LiftLeaseComparison shape via the API and set the value
            txt_LiftLease = sheet_dashboard.shapes['txtDLLeaseDiff'].api
            txt_LiftLease.TextFrame2.TextRange.Text = f"${prev_LeaseComp} to ${lat_LeaseComp}"
            txt_LiftLease_diff = sheet_dashboard.shapes['txtLLeaseDiff'].api
            txt_LiftLease_diff.TextFrame2.TextRange.Text = f"${diff_LeaseComp}"

            # Access the ViolationComparison shape via the API and set the value
            txt_Violation = sheet_dashboard.shapes['txtDViolationsDiff'].api
            txt_Violation.TextFrame2.TextRange.Text = f"${prev_ViolationComp} to ${lat_ViolationComp}"
            txt_Violation_diff = sheet_dashboard.shapes['txtViolationsDiff'].api
            txt_Violation_diff.TextFrame2.TextRange.Text = f"${diff_ViolationComp}"

            # Access the OperatorChanges shape via the API and set the value
            # txt_Operator = sheet_dashboard.shapes['txtDOperatorsDiff'].api
            # txt_Operator.TextFrame2.TextRange.Text = f"{prev_OperatorChanges} to {lat_OperatorChanges} operators"
            # txt_Operator_diff = sheet_dashboard.shapes['txtOperatorsDiff'].api
            # txt_Operator_diff.TextFrame2.TextRange.Text = f"{diff_OperatorChanges} operators"

            # Run the VBA macro to update the color based on the status
            try:
                # Parameters: TextBox name and status
                textBoxName = "TextBox 90"
                status = full_amount_diff_status  # Use the status from the comparison file

                # Call the VBA macro to update color
                wb_dashboard.macro("UpdateTextBoxColor")(sheet_name, textBoxName, status)
                logger.info(f"Successfully updated color for {textBoxName} with status '{status}'.")
            except Exception as e:
                logger.error(f"An error occurred: {e}")

            # Run the VBA macro to update the color based on the values
            try:
                # Parameters: TextBox names and corresponding values
                # textBoxNames = ["txtHTotalRevDiff", "txtLLeaseDiff", "txtViolationsDiff", "txtOperatorsDiff"]
                # values = [diff_HTotalRev, diff_LeaseComp, diff_ViolationComp, diff_OperatorChanges]

                textBoxNames = ["txtHTotalRevDiff", "txtLLeaseDiff", "txtViolationsDiff"]
                values = [diff_HTotalRev, diff_LeaseComp, diff_ViolationComp]

                # Loop through the text boxes and update colors based on the values
                for i, textBoxName in enumerate(textBoxNames):
                    wb_dashboard.macro("UpdateSummaryColor")(sheet_name, textBoxName, values[i])
                    logger.info(f"Successfully updated color for {textBoxName} with value '{values[i]}'.")
            except Exception as e:
                logger.error(f"A TextBox error occurred: {e}")

        # Save the changes to the dashboard workbook
        wb_dashboard.save()
        wb_dashboard.close()
        logger.info(f"{dashboard_file} has been successfully updated and saved.")

        # paste_picture(comparison_files, dashboard_file)
        app.quit()
        time.sleep(2)
        paste_picture()
        time.sleep(2)
        wt.main(file_previous, file_latest)

    except Exception as e:
        logger.info(f"A Dashboard error occurred: {e}")
    # finally:
    #     wb_dashboard.save()
    #     wb_dashboard.close()
        # app.quit()
        
        # Reopen the Excel file
        # app = xw.App(visible=True)  # Open Excel with the app visible
        # wb_dashboard = app.books.open(dashboard_file)  # Reopen the file

def paste_picture():
    comparison_files = [
        ('ComparedResults\\DIV2_Tables.xlsx', 'Dashboard')
    ]
    
    # Target cells for each sheet in the comparison file
    target_cells = {
        'ViolationComparison': (30, 12),
        'HTotalRevComparison': (14, 12),
        # 'OperatorChanges': (22, 19),
        'LiftLeaseComparison': (5, 19)
    }

    relative_dashboard_path = "ComparedResults\\Dashboard.xlsm"
    # Get the absolute path of the current script's directory
    # script_dir = os.path.dirname(os.path.realpath(__file__))

    # # Build the full path to the dashboard file by joining the script directory and the relative path
    # dashboard_file = os.path.join(script_dir, relative_dashboard_path)

    # Get the base directory correctly whether running as script or PyInstaller executable
    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)  # Executable folder
    else:
        script_dir = os.path.dirname(os.path.realpath(__file__))  # Script folder

    dashboard_file = os.path.join(script_dir, "ComparedResults", "Dashboard.xlsm")

    excel = None

    try:
        # Initialize Excel application
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True  # Set to True for debugging

        # Check if the file exists
        if not os.path.exists(dashboard_file):
            logger.info(f"Error: Dashboard file does not exist at {dashboard_file}")
            return
        
        # Open the Dashboard workbook
        wb_dashboard = excel.Workbooks.Open(dashboard_file)
        if wb_dashboard is None:
            logger.info(f"Failed to open the Dashboard workbook at {dashboard_file}")
            return

        # Delete existing pictures if they exist
        for target_sheet_name in ['Dashboard']:
            ws_dashboard = wb_dashboard.Sheets(target_sheet_name)
            ws_dashboard.Activate()
            # for picture_name in ['ViolationsTable', 'HoursTable', 'OperatorTable', 'LeaseTable']:
            for picture_name in ['ViolationsTable', 'HoursTable', 'LeaseTable']:
                try:
                    ws_dashboard.Shapes(picture_name).Delete()  # Attempt to delete the picture
                    logger.info(f"Deleted existing picture: {picture_name} in {target_sheet_name}")
                except Exception:
                    logger.info(f"No existing picture named {picture_name} found in {target_sheet_name}")  

        time.sleep(2)
        
        # Process each comparison file
        for comparison_file, target_sheet_name in comparison_files:
            # Build the full path for the comparison file
            comparison_file_path = os.path.join(script_dir, comparison_file)
            
            # Check if the comparison file exists
            if not os.path.exists(comparison_file_path):
                logger.info(f"Error: Comparison file does not exist at {comparison_file_path}")
                continue

            # Open the comparison workbook
            wb_comparison = excel.Workbooks.Open(comparison_file_path)
            if wb_comparison is None:
                logger.info(f"Failed to open the comparison workbook at {comparison_file_path}")
                continue

            # Process each sheet in the comparison file (ViolationComparison, HTotalRevComparison, OperatorChanges)
            for sheet_name, target_cell in target_cells.items():
                sheet = wb_comparison.Sheets(sheet_name)
                table_width = 0
                table_height = 0
                table_name = None
                if sheet is None:
                    logger.info(f"Failed to access the '{sheet_name}' sheet in {comparison_file_path}")
                    continue

                # Get the used range
                used_range = sheet.UsedRange

                time.sleep(1)
                if used_range.Rows.Count > 1 and used_range.Columns.Count > 1:
                    try:
                        table_width = (used_range.Width) # Get the width of the used range
                        table_height = (used_range.Height) # Get the height of the used range
                        used_range.CopyPicture(Format=2)
                        logger.info(f"Copied picture from {sheet_name}, width: {table_width*0.0352778:.2f} cm, height: {table_height*0.0352778:.2f} cm")
                    except Exception as e:
                        logger.error(f"Failed to copy picture from {sheet_name}: {e}")
                        continue
                else:
                    logger.error(f"Skipping {sheet_name}: No data in the range.")
                    continue  # Skip further processing for this sheet

                # Activate the target sheet in the Dashboard workbook
                ws_dashboard = wb_dashboard.Sheets(target_sheet_name)
                if ws_dashboard is None:
                    logger.info(f"Failed to access the sheet '{target_sheet_name}' in the Dashboard workbook.")
                    wb_comparison.Close(SaveChanges=False)
                    continue
                
                time.sleep(1)
                # Paste as a picture into the target sheet
                ws_dashboard.Activate()
                row, col = target_cell
                target_cell_range = ws_dashboard.Cells(row, col)  # Adjust as needed
                ws_dashboard.Paste(target_cell_range)

                # Position and resize the pasted picture
                pasted_picture = ws_dashboard.Shapes(ws_dashboard.Shapes.Count)
                pasted_picture.Left = target_cell_range.Left
                pasted_picture.Top = target_cell_range.Top

                # Name the pasted picture according to the sheet
                if sheet_name == 'ViolationComparison':
                    pasted_picture.Name = 'ViolationsTable'
                elif sheet_name == 'HTotalRevComparison':
                    pasted_picture.Name = 'HoursTable'
                # elif sheet_name == 'OperatorChanges':
                #     pasted_picture.Name = 'OperatorTable'
                elif sheet_name == 'LiftLeaseComparison':
                    pasted_picture.Name = 'LeaseTable'

                table_name = pasted_picture.Name
                logger.info(f"Table Name: {table_name}")

                time.sleep(1)
                logger.info(f"'{pasted_picture.Name}' successfully pasted in the {sheet_name} Sheet.")

                # Adjust the container size based on the table size
                # Adjust only the container that matches the table name
                if table_name:
                    container_name = table_name.replace("Table", "Container")  # Match the container name
                    try:
                        container = ws_dashboard.Shapes(container_name)
                        # container.Width = table_width + 95  # Add 3.35 cm to width
                        container.Width = table_width # Add 3.35 cm to width
                        # container.Height = table_height + 123  # Add 4.33 cm to height
                        container.Height = table_height + 56  # Add 4.33 cm to height
                        logger.info(f"Resized {container_name} to width: {(container.Width)*0.0352778:.2f} cm, height: {(container.Height)*0.0352778:.2f} cm")
                    except Exception as e:
                        logger.error(f"Failed to resize {container_name}: {e}")

                # wb_comparison.Save()
                # wb_dashboard.Save()
        wb_comparison.Save()
        wb_comparison.Close() 
            # Close the comparison workbook without saving
            # wb_comparison.Close(SaveChanges=True)

        # Save and close the Dashboard workbook
        wb_dashboard.Save()
        wb_dashboard.Close()
        excel.Quit()

        logger.info("Data pasted as pictures successfully.")

        app = xw.App(visible=True)  # Open Excel with the app visible
        wb_dashboard = app.books.open(dashboard_file)  # Reopen the file
    except Exception as e:
        logger.error(f"An error occurred: {e}")

        if 'wb_dashboard' in locals() and wb_dashboard:
            wb_dashboard.Close(SaveChanges=False)

        if excel:
            excel.Quit()
            del excel
        # logger.error(f"An error occurred: {e}")
        # wb_dashboard.Close()
        # excel.Quit()
        # del excel 
    finally:
        # Ensure Excel is properly quit and the object is released
        if excel:
            excel.Quit()  # Close the Excel application
            del excel 
        # app = xw.App(visible=True)  # Open Excel with the app visible
        # wb_dashboard = app.books.open(dashboard_file)  # Reopen the file

# def adjust_container(table_name, table_width, table_height):
#     for target_sheet_name in ['Dashboard', 'CCCTA', 'LAVTA']:
#             ws_dashboard = wb_dashboard.Sheets(target_sheet_name)
#             ws_dashboard.Activate()
#             for picture_name in ['TripsTable', 'HoursTable', 'OperatorTable', 'LeaseTable']:
#                 try:
#                     ws_dashboard.Shapes(picture_name).Delete()  # Attempt to delete the picture
#                     logger.info(f"Deleted existing picture: {picture_name} in {target_sheet_name}")
#                 except Exception:
#                     logger.info(f"No existing picture named {picture_name} found in {target_sheet_name}")  

#     # Get the active table and container dimensions
#     if table_name == 'TripsTable':
#     elif table_name == 'HoursTable':
#     elif table_name == 'OperatorTable':
#     elif table_name == 'LeaseTable':

#     # Calculate the aspect ratios
#     table_aspect_ratio = table_width / table_height
#     container_aspect_ratio = container_width / container_height

#     # Determine the scaling factor
#     if table_aspect_ratio > container_aspect_ratio:
#         # Scale based on width
#         scale_factor = container_width / table_width
#     else:
#         # Scale based on height
#         scale_factor = container_height / table_height

#     # Calculate the new dimensions
#     new_width = table_width * scale_factor
#     new_height = table_height * scale_factor

#     return new_width, new_height

# if __name__ == '__main__':
    
#     comparison_files = [
#         ('Compared Results/Full_Comparison.xlsx', 'Dashboard'),
#         ('Compared Results/CCCTA_Comparison.xlsx', 'CCCTA'),
#         ('Compared Results/LAVTA_Comparison.xlsx', 'LAVTA')
#     ]
#     dashboard_file = 'Compared Results/Dashboard.xlsm'
#     main(file_previous, file_latest)
#     paste_picture(comparison_files, dashboard_file)
#     paste_picture(comparison_files, dashboard_file)
