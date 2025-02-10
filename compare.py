import pandas as pd
import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import tkinter as tk
from tkinter import filedialog
import dashboard as db
import time

log_folder = "Logs"
os.makedirs(log_folder, exist_ok=True)
# Set up logging
logging.basicConfig(
    level=logging.DEBUG, 
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('Logs/Comparison.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger()

# Declare global variables
file_entry_previous = None
file_entry_latest = None

calculated_amount = []
# calculated_totals = 0

def load_sheets(file_previous, file_latest):
    try:
        logger.info("Loading sheets from the provided Excel files.")
        sheet_pr_previous = pd.read_excel(file_previous, sheet_name="PR")
        sheet_pr_latest = pd.read_excel(file_latest, sheet_name="PR")

        sheet_vdpmv_previous = pd.read_excel(file_previous, sheet_name="SumVDPMVReport")
        sheet_vdpmv_latest = pd.read_excel(file_latest, sheet_name="SumVDPMVReport")

        sheet_partner_previous = pd.read_excel(file_previous, sheet_name="Div2PartnerList")
        sheet_partner_latest = pd.read_excel(file_latest, sheet_name="Div2PartnerList")

        sheet_deductions_previous = pd.read_excel(file_previous, sheet_name="Deduction and other Revenue")
        sheet_deductions_latest = pd.read_excel(file_latest, sheet_name="Deduction and other Revenue")
        
        logger.info("Sheets loaded successfully.")
        return sheet_pr_previous, sheet_pr_latest, sheet_vdpmv_previous, sheet_vdpmv_latest, sheet_partner_previous, sheet_partner_latest, sheet_deductions_previous, sheet_deductions_latest
    except Exception as e:
        logger.error(f"Error loading sheets: {e}")
        raise

def clean_currency(value):
    try:
        if isinstance(value, str):
            value = value.replace('$', '').replace(',', '').strip()
            return round(float(value), 2) if value else None
        return value
    except ValueError:
        logger.error(f"Error cleaning currency value: {value}")
        return None

def calculate_totals(deductions_sheet, pr_sheet):
    try:
        calculated_totals = 0
        clients = deductions_sheet["TYPE"].unique()
        print("Clients in deductions_sheet:", clients)
        for client in clients: 
            # logger.info(f"Calculating totals for {client} clients.")

            client_header_row = pr_sheet[pr_sheet.iloc[:, 0].astype(str).str.contains(client, na=False, case=False)]
            if client_header_row.empty:
                logger.warning(f"No header row found for client {client} in pr_sheet.")
                continue  # Skip to the next client
            # Get the index of the client header row
            client_header_index = client_header_row.index[0]
            # logger.info(f"Found client header for {client} at row {client_header_index}.")

            # Find the first empty row after the client header row to determine the endpoint
            empty_row_index = pr_sheet.iloc[client_header_index + 1:, 0].isna().idxmax()

            if pd.isna(empty_row_index):
                # If no empty row is found, use the last row of the DataFrame
                next_header_index = pr_sheet.shape[0]
            else:
                # The index of the first empty row marks the endpoint
                next_header_index = empty_row_index + client_header_index + 1

            # logger.info(f"Partner rows for client {client} are between rows {client_header_index + 1} and {next_header_index - 1}.")

            # Extract the partner rows between the client header and the first empty row
            partner_rows = pr_sheet.iloc[client_header_index + 1:next_header_index]

            # Get all unique partners for this client from the deductions sheet
            total_partners = deductions_sheet["PARTNER"].unique()
            # logger.info(f"Total partners for {client}: {len(total_partners)}")

            total_amount = 0

            # Calculate total amount for matching partners
            for partner in total_partners:
                # Find rows in partner_rows matching the partner
                partner_rows_matched = partner_rows[partner_rows.iloc[:, 0].astype(str).str.strip() == str(partner).strip()]

                if not partner_rows_matched.empty:
                    # Add the amount found in column 16 (assuming it's the amount column)
                    total_amount += partner_rows_matched.iloc[0, 16]  # Column 16 contains the amount
                    # logger.info(f"Amount for partner {partner}: {partner_rows_matched.iloc[0, 14]}")
            calculated_totals += total_amount
        logger.info(f"Total amount for client {client}: {total_amount}")
        # logger.info("Calculating totals for trips, hours, operators, and amounts.")
        logger.info(f"Total calculated mamount:{calculated_totals}")
        return calculated_totals
    except Exception as e:
        logger.error(f"Error calculating totals: {e}")
        raise

def compare_totals(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing Totals between previous and latest values.")

        # Ensure inputs are numeric
        if not isinstance(sheet_previous, (int, float)) or not isinstance(sheet_latest, (int, float)):
            raise TypeError("Both sheet_previous and sheet_latest must be numeric values.")
        
        # Create a DataFrame to store results
        deductions_comparison = pd.DataFrame({
            "LATEST": [sheet_latest],
            "PREVIOUS": [sheet_previous],
            "DIFFERENCE": [sheet_latest - sheet_previous]
        }).round(2)

        # Add CHANGE column based on the difference
        deductions_comparison["CHANGE"] = deductions_comparison["DIFFERENCE"].apply(
            lambda diff: "Increased" if diff > 0 else "Decreased" if diff < 0 else "No Change"
        )

        logger.info("Totals comparison completed.")
        return deductions_comparison

    except Exception as e:
        logger.error(f"Error comparing totals: {e}")
        raise

def compare_htotalrev(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing Hourly Total Revs between previous and latest sheets.")
        
        # Select the relevant columns directly
        previous_values = sheet_previous[["PARTNER NAME", "Total Rev"]]
        latest_values = sheet_latest[["PARTNER NAME", "Total Rev"]]

        # Merge both dataframes on "PARTNER" and handle missing values with 0
        comparison = previous_values.merge(latest_values, on="PARTNER NAME", how="outer", suffixes=("_PREVIOUS", "_LATEST")).fillna(0)

        # Calculate the change in the "Total Rev"
        comparison["CHANGE"] = comparison["Total Rev_LATEST"] - comparison["Total Rev_PREVIOUS"]
        for col in ["Total Rev_LATEST", "Total Rev_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Prepare the final dataframe for comparison
        deductions_comparison = comparison[["PARTNER NAME", "Total Rev_LATEST", "Total Rev_PREVIOUS", "CHANGE"]].drop_duplicates(subset="PARTNER NAME")

        # Rename columns
        deductions_comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("HTOTALREV comparison completed.")
        return deductions_comparison

    except Exception as e:
        logger.error(f"Error comparing HTOTALREV: {e}")
        raise

def compare_liftlease(sheet_previous, sheet_latest, htotalrev_df):
    try:
        logger.info("Comparing Lift Lease between previous and latest sheets.")
        
        # Group by PARTNER and sum the LIFT LEASE TOTAL
        previous_values = sheet_previous.groupby("PARTNER", as_index=False)["LIFT LEASE TOTAL"].sum()
        latest_values = sheet_latest.groupby("PARTNER", as_index=False)["LIFT LEASE TOTAL"].sum()


        # Merge both dataframes on "PARTNER" and handle missing values with 0
        comparison = previous_values.merge(latest_values, on="PARTNER", how="outer", suffixes=("_PREVIOUS", "_LATEST")).fillna(0)
        comparison = comparison.merge(htotalrev_df, on="PARTNER", how="inner").fillna(0)

        # Calculate the change in the "LIFT LEASE TOTAL"
        comparison["CHANGE"] = comparison["LIFT LEASE TOTAL_LATEST"] - comparison["LIFT LEASE TOTAL_PREVIOUS"]
        for col in ["LIFT LEASE TOTAL_LATEST", "LIFT LEASE TOTAL_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Prepare the final dataframe for comparison
        deductions_comparison = comparison[["PARTNER", "LIFT LEASE TOTAL_LATEST", "LIFT LEASE TOTAL_PREVIOUS", "CHANGE"]].drop_duplicates(subset="PARTNER")

        # Rename columns
        deductions_comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("Lift Lease comparison completed.")
        return deductions_comparison

    except Exception as e:
        logger.error(f"Error comparing Lift Lease: {e}")
        raise

def compare_violations(sheet_previous, sheet_latest, htotalrev_df):
    try:
        logger.info("Comparing Lift Lease between previous and latest sheets.")
        
        # Group by PARTNER and sum the Violation
        previous_values = sheet_previous.groupby("PARTNER", as_index=False)["Violation"].sum()
        latest_values = sheet_latest.groupby("PARTNER", as_index=False)["Violation"].sum()


        # Merge both dataframes on "PARTNER" and handle missing values with 0
        comparison = previous_values.merge(latest_values, on="PARTNER", how="outer", suffixes=("_PREVIOUS", "_LATEST")).fillna(0)
        comparison = comparison.merge(htotalrev_df, on="PARTNER", how="inner").fillna(0)

        # Calculate the change in the "LIFT LEASE TOTAL"
        comparison["CHANGE"] = comparison["Violation_LATEST"] - comparison["Violation_PREVIOUS"]

        for col in ["Violation_LATEST", "Violation_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].round(2)

        # Prepare the final dataframe for comparison
        deductions_comparison = comparison[["PARTNER", "Violation_LATEST", "Violation_PREVIOUS", "CHANGE"]].drop_duplicates(subset="PARTNER")

        # Rename columns
        deductions_comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info("Violation comparison completed.")
        return deductions_comparison

    except Exception as e:
        logger.error(f"Error comparing Violation: {e}")
        raise

def compare_operators(sheet_previous, sheet_latest):
    # try:
    #     logger.info("Comparing operators between previous and latest sheets.")
        
    #     # Extract and store unique pairs of (OPERATOR NAME, PARTNER NAME)
    #     operators_previous = set(sheet_previous[["OPERATOR NAME", "PARTNER NAME"]].dropna().itertuples(index=False, name=None))
    #     operators_latest = set(sheet_latest[["OPERATOR NAME", "PARTNER NAME"]].dropna().itertuples(index=False, name=None))

    #     # Identify added and removed operators
    #     added = operators_latest - operators_previous
    #     removed = operators_previous - operators_latest

    #     # Convert to DataFrame format
    #     added_list = [{"Operator Name": op, "Partner": partner, "Change": "Added"} for op, partner in added]
    #     removed_list = [{"Operator Name": op, "Partner": partner, "Change": "Removed"} for op, partner in removed]

    #     # Create DataFrames
    #     added_df = pd.DataFrame(added_list)
    #     removed_df = pd.DataFrame(removed_list)

    #     # Ensure DataFrames always have the necessary columns
    #     if added_df.empty:
    #         added_df = pd.DataFrame(columns=["Operator Name", "Partner", "Change"])
    #     if removed_df.empty:
    #         removed_df = pd.DataFrame(columns=["Operator Name", "Partner", "Change"])

    #     # Combine the results
    #     operator_changes_df = pd.concat([added_df, removed_df], ignore_index=True)

    #     # If no changes were found, add a "No Changes" row
    #     if operator_changes_df.empty:
    #         operator_changes_df = pd.DataFrame([{
    #             "Operator Name": "No Changes",
    #             "Partner": "No Changes",
    #             "Change": "No Changes"
    #         }])

    #     logger.info("Operator comparison completed.")
    #     return operator_changes_df

    # except Exception as e:
    #     logger.error(f"Error comparing operators: {e}")
    #     raise
    try:
        logger.info("Comparing operators between previous and latest sheets.")
        
        # Extract and store unique pairs of (OPERATOR NAME, PARTNER NAME)
        operators_previous = set(sheet_previous[["OPERATOR NAME", "PARTNER NAME"]].dropna().itertuples(index=False, name=None))
        operators_latest = set(sheet_latest[["OPERATOR NAME", "PARTNER NAME"]].dropna().itertuples(index=False, name=None))

        # Identify added and removed operators
        added = operators_latest - operators_previous
        removed = operators_previous - operators_latest

        # Convert to DataFrame format
        added_list = [{"Operator Name": op, "Partner": partner, "Change": "Added"} for op, partner in added]
        removed_list = [{"Operator Name": op, "Partner": partner, "Change": "Removed"} for op, partner in removed]

        # Create DataFrames
        added_df = pd.DataFrame(added_list)
        removed_df = pd.DataFrame(removed_list)

        # Ensure DataFrames always have the necessary columns
        if added_df.empty:
            added_df = pd.DataFrame(columns=["Operator Name", "Partner", "Change"])
        if removed_df.empty:
            removed_df = pd.DataFrame(columns=["Operator Name", "Partner", "Change"])

        # Combine the results
        operator_changes_df = pd.concat([added_df, removed_df], ignore_index=True)

        logger.info("Operator comparison completed.")
        return operator_changes_df

    except Exception as e:
        logger.error(f"Error comparing operators: {e}")
        raise

def compare_acceptance_rate(sheet_previous, sheet_latest, week):
    try:
        logger.info(f"Comparing Acceptance Rate for {week}.")

        # Filter data for the specific week
        prev_week = sheet_previous[sheet_previous["WeekN"] == week]
        latest_week = sheet_latest[sheet_latest["WeekN"] == week]

        # Group by "PARTNER NAME" and sum Acceptance Rate
        prev_values = prev_week.groupby("PARTNER NAME", as_index=False)["Acceptance Rate"].sum()
        latest_values = latest_week.groupby("PARTNER NAME", as_index=False)["Acceptance Rate"].sum()

        # Merge both datasets
        comparison = prev_values.merge(
            latest_values, on="PARTNER NAME", how="outer", suffixes=("_PREVIOUS", "_LATEST")
        ).fillna(0)

        # Calculate the change
        comparison["CHANGE"] = comparison["Acceptance Rate_LATEST"] - comparison["Acceptance Rate_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["Acceptance Rate_LATEST", "Acceptance Rate_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].apply(lambda x: float(f"{x * 100:.2f}"))

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info(f"{week} Acceptance Rate comparison completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Acceptance Rate for {week}: {e}")
        raise


def compare_cancellation_rate(sheet_previous, sheet_latest, week):
    try:
        logger.info(f"Comparing Cancellation Rate for {week}.")

        # Filter data for the specific week
        prev_week = sheet_previous[sheet_previous["WeekN"] == week]
        latest_week = sheet_latest[sheet_latest["WeekN"] == week]

        # Group by "PARTNER NAME" and sum Cancellation Rate
        prev_values = prev_week.groupby("PARTNER NAME", as_index=False)["Cancellation Rate"].sum()
        latest_values = latest_week.groupby("PARTNER NAME", as_index=False)["Cancellation Rate"].sum()

        # Merge both datasets
        comparison = prev_values.merge(
            latest_values, on="PARTNER NAME", how="outer", suffixes=("_PREVIOUS", "_LATEST")
        ).fillna(0)

        # Calculate the change
        comparison["CHANGE"] = comparison["Cancellation Rate_LATEST"] - comparison["Cancellation Rate_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["Cancellation Rate_LATEST", "Cancellation Rate_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].apply(lambda x: float(f"{x * 100:.2f}"))

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info(f"{week} Cancellation Rate comparison completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Cancellation Rate for {week}: {e}")
        raise


def compare_utilization(sheet_previous, sheet_latest, week):
    try:
        logger.info(f"Comparing Utilization% for {week}.")

        # Filter data for the specific week
        prev_week = sheet_previous[sheet_previous["WeekN"] == week]
        latest_week = sheet_latest[sheet_latest["WeekN"] == week]

        # Group by "PARTNER NAME" and sum Utilization%
        prev_values = prev_week.groupby("PARTNER NAME", as_index=False)["Utilization%"].sum()
        latest_values = latest_week.groupby("PARTNER NAME", as_index=False)["Utilization%"].sum()

        # Merge both datasets
        comparison = prev_values.merge(
            latest_values, on="PARTNER NAME", how="outer", suffixes=("_PREVIOUS", "_LATEST")
        ).fillna(0)

        # Calculate the change
        comparison["CHANGE"] = comparison["Utilization%_LATEST"] - comparison["Utilization%_PREVIOUS"]

        # Format columns as percentages with 4 significant digits
        for col in ["Utilization%_LATEST", "Utilization%_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].apply(lambda x: float(f"{x * 100:.2f}"))

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info(f"{week} Utilization% comparison completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Utilization% for {week}: {e}")
        raise


def compare_pNormalHours(sheet_previous, sheet_latest, week):
    try:
        logger.info(f"Comparing Payable Normal Hours for {week}.")

        # Filter data for the specific week
        prev_week = sheet_previous[sheet_previous["WeekN"] == week]
        latest_week = sheet_latest[sheet_latest["WeekN"] == week]

        # Group by "PARTNER NAME" and sum Payable Normal Hours
        prev_values = prev_week.groupby("PARTNER NAME", as_index=False)["Payable Normal Hours"].sum()
        latest_values = latest_week.groupby("PARTNER NAME", as_index=False)["Payable Normal Hours"].sum()

        # Merge both datasets
        comparison = prev_values.merge(
            latest_values, on="PARTNER NAME", how="outer", suffixes=("_PREVIOUS", "_LATEST")
        ).fillna(0)

        # Calculate the change
        comparison["CHANGE"] = comparison["Payable Normal Hours_LATEST"] - comparison["Payable Normal Hours_PREVIOUS"]
        for col in ["Payable Normal Hours_LATEST", "Payable Normal Hours_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].apply(lambda x: float(f"{x:.2f}"))

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info(f"{week} Payable Normal Hours comparison completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Payable Normal Hours for {week}: {e}")
        raise

def compare_pBonusHours(sheet_previous, sheet_latest, week):
    try:
        logger.info(f"Comparing Payable Bonus Hours for {week}.")

        # Filter data for the specific week
        prev_week = sheet_previous[sheet_previous["WeekN"] == week]
        latest_week = sheet_latest[sheet_latest["WeekN"] == week]

        # Group by "PARTNER NAME" and sum Payable Bonus Hours
        prev_values = prev_week.groupby("PARTNER NAME", as_index=False)["Payable Bonus Hours"].sum()
        latest_values = latest_week.groupby("PARTNER NAME", as_index=False)["Payable Bonus Hours"].sum()

        # Merge both datasets
        comparison = prev_values.merge(
            latest_values, on="PARTNER NAME", how="outer", suffixes=("_PREVIOUS", "_LATEST")
        ).fillna(0)

        # Calculate the change
        comparison["CHANGE"] = comparison["Payable Bonus Hours_LATEST"] - comparison["Payable Bonus Hours_PREVIOUS"]
        for col in ["Payable Bonus Hours_LATEST", "Payable Bonus Hours_PREVIOUS", "CHANGE"]:
            comparison[col] = comparison[col].apply(lambda x: float(f"{x:.2f}"))

        # Rename columns
        comparison.columns = ["PARTNER", "LATEST", "PREVIOUS", "CHANGE"]

        logger.info(f"{week} Payable Bonus Hours comparison completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Payable Bonus Hours for {week}: {e}")
        raise

def apply_formatting(sheet_name, wb):
    try:
        logger.info(f"Applying formatting to sheet: {sheet_name}.")
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                if ws[1][cell.column - 1].value.lower() == "change":
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100")
                        elif cell.value < 0:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006")
                    elif isinstance(cell.value, str):
                        if cell.value.lower() in ["increased", "added"]:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100")
                        elif cell.value.lower() in ["decreased", "removed"]:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006")
        logger.info(f"Formatting applied successfully to sheet: {sheet_name}.")
    except Exception as e:
        logger.error(f"Error applying formatting to sheet {sheet_name}: {e}")
        raise


def save_comparison_results(output_folder, comparison_data, filename):
    try:
        logger.info(f"Saving comparison results to {filename}.")
        os.makedirs(output_folder, exist_ok=True)
        full_comparison_file = os.path.join(output_folder, filename)
        with pd.ExcelWriter(full_comparison_file, engine="openpyxl") as writer:
            for sheet_name, data in comparison_data.items():
                data.to_excel(writer, sheet_name=sheet_name, index=False)

        wb_full = load_workbook(full_comparison_file)
        for sheet in comparison_data.keys():
            apply_formatting(sheet, wb_full)
        wb_full.save(full_comparison_file)
        wb_full.close()

        logger.info(f"Comparison results saved successfully to {filename}.")
    except Exception as e:
        logger.error(f"Error saving comparison results: {e}")
        raise


def main(file_previous, file_latest):
    try: 
        file_entry_previous = file_previous
        file_entry_latest = file_latest

        print(f"{file_entry_previous} + {file_entry_latest}")
        logger.info("Starting main comparison process.")
        output_folder = "ComparedResults"
        os.makedirs(output_folder, exist_ok=True)

        sheet_pr_previous, sheet_pr_latest, sheet_vdpmv_previous, sheet_vdpmv_latest, sheet_partner_previous, sheet_partner_latest, sheet_deductions_previous, sheet_deductions_latest = load_sheets(file_previous, file_latest)

        # 1. Process the data without any filtering (full comparison)

        # Compare totals
        prev_totals = calculate_totals(sheet_deductions_previous, sheet_pr_previous)
        lat_totals = calculate_totals(sheet_deductions_latest, sheet_pr_latest)
        totals_comparison_df = compare_totals(prev_totals, lat_totals)
        # logger.info(f"Totals DF: {totals_comparison_df}")

        # 2. Compare hourly total revs
        compare_htotalrev_df = compare_htotalrev(sheet_vdpmv_previous, sheet_vdpmv_latest)
        # logger.info(f"HTOTAL_REV DF: {compare_htotalrev_df}")

        # 3. Compare Lift Lease
        compare_liftlease_df = compare_liftlease(sheet_deductions_previous, sheet_deductions_latest, compare_htotalrev_df)
        # logger.info(f"Lift Lease DF: {compare_liftlease_df}")

        # 4. Compare Violations
        compare_violations_df = compare_violations(sheet_deductions_previous, sheet_deductions_latest, compare_htotalrev_df)
        # logger.info(f"Violations DF: {compare_violations_df}")

        # 5. Compare operators
        operator_changes_df = compare_operators(sheet_partner_previous, sheet_partner_latest)
        # logger.info(f"Operator Changes DF:{operator_changes_df}")

        # 6. Week 1 comparison
        

        # Save the full comparison results
        full_comparison_file = os.path.join(output_folder, "DIV2_Tables.xlsx")
        excel_sheets = []
        with pd.ExcelWriter(full_comparison_file, engine="openpyxl") as writer:
            totals_comparison_df.to_excel(writer, sheet_name="TotalInvoicePayment", index=False)
            excel_sheets.append("TotalInvoicePayment")
            compare_htotalrev_df.to_excel(writer, sheet_name="HTotalRevComparison", index=False)
            excel_sheets.append("HTotalRevComparison")
            compare_liftlease_df.to_excel(writer, sheet_name="LiftLeaseComparison", index=False)
            excel_sheets.append("LiftLeaseComparison")
            compare_violations_df.to_excel(writer, sheet_name="ViolationComparison", index=False)
            excel_sheets.append("ViolationComparison")
            operator_changes_df.to_excel(writer, sheet_name="OperatorChanges", index=False)
            excel_sheets.append("OperatorChanges")
            for week in ["Week1", "Week2"]:
                acceptance_changes_df = compare_acceptance_rate(sheet_vdpmv_previous, sheet_vdpmv_latest, week)
                acceptance_changes_df.to_excel(writer, sheet_name=f"{week}AcceptRateComp", index=False)
                excel_sheets.append(f"{week}AcceptRateComp")

                cancellation_changes_df = compare_cancellation_rate(sheet_vdpmv_previous, sheet_vdpmv_latest, week)
                cancellation_changes_df.to_excel(writer, sheet_name=f"{week}CancelRateComp", index=False)
                excel_sheets.append(f"{week}CancelRateComp")

                utilization_changes_df = compare_utilization(sheet_vdpmv_previous, sheet_vdpmv_latest, week)
                utilization_changes_df.to_excel(writer, sheet_name=f"{week}UtilizationComp", index=False)
                excel_sheets.append(f"{week}UtilizationComp")

                pnormalhours_changes_df = compare_pNormalHours(sheet_vdpmv_previous, sheet_vdpmv_latest, week)
                pnormalhours_changes_df.to_excel(writer, sheet_name=f"{week}PNormalHrsComp", index=False)
                excel_sheets.append(f"{week}PNormalHrsComp")

                pbonushours_changes_df = compare_pBonusHours(sheet_vdpmv_previous, sheet_vdpmv_latest, week)
                pbonushours_changes_df.to_excel(writer, sheet_name=f"{week}PBonusHrsComp", index=False)
                excel_sheets.append(f"{week}PBonusHrsComp")

        # Doperator_changes_df.to_excel(writer, sheet_name="DateComparison", index=False)

        # Apply formatting to the full comparison file
        wb_full = load_workbook(full_comparison_file)
        for sheet in excel_sheets:                                                                                                
            apply_formatting(sheet, wb_full)
        wb_full.save(full_comparison_file)
        wb_full.close()

        logger.info(f"Main comparison process completed successfully. File saved to {full_comparison_file}.")
        # time.sleep(2)
        db.main(file_previous, file_latest)
    except Exception as e:
        logger.error(f"Error in main comparison process: {e}")
        raise
    # finally:
    #     wb_client.close()


def open_file_dialog(entry):
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsm;*.xlsx")])
    if filename:
        entry.delete(0, tk.END)
        entry.insert(0, filename)


def create_gui():
    global file_entry_previous, file_entry_latest # Declare them as global variables

    # Set up the GUI window
    root = tk.Tk()
    root.title("Comparison Tool")

    # Create and place labels, entry boxes, and buttons
    tk.Label(root, text="Previous File:").grid(row=0, column=0, padx=10, pady=5)
    entry_previous = tk.Entry(root, width=50)
    entry_previous.grid(row=0, column=1, padx=10, pady=5)
    
    # tk.Button(root, text="Browse", command=lambda: open_file_dialog(entry_previous)).grid(row=0, column=2, padx=10, pady=5)

    tk.Label(root, text="Latest File:").grid(row=1, column=0, padx=10, pady=5)
    entry_latest = tk.Entry(root, width=50)
    entry_latest.grid(row=1, column=1, padx=10, pady=5)
    
    # tk.Button(root, text="Browse", command=lambda: open_file_dialog(entry_latest)).grid(row=1, column=2, padx=10, pady=5)

    # Button to trigger the comparison process
    # tk.Button(root, text="Compare", command=lambda: (main(entry_previous.get(), entry_latest.get()), root.destroy())).grid(row=2, column=1, pady=20)
    tk.Button(root, text="Compare", command=lambda: handle_comparison(entry_previous.get(), entry_latest.get(), root)).grid(row=2, column=1, pady=20)

    def handle_comparison(file_previous, file_latest, root):
        try:
            main(file_previous, file_latest)
        except Exception as e:
            print(f"An error occurred: {e}")
            # Check for the disconnection error and close the GUI if it happens
            if isinstance(e, OSError) and "The object invoked has disconnected" in str(e):
                print("Disconnected from Excel, closing GUI.")
                root.quit()  # This will close the Tkinter window
        finally:
            root.destroy()  # Close the window in all cases
    # Start the GUI loop
    root.mainloop()


if __name__ == "__main__":
    create_gui()

