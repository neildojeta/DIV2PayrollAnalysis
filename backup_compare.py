import pandas as pd
import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import tkinter as tk
from tkinter import filedialog
import dashboard as db
import time

log_folder = "Logs"
os.makedirs(log_folder, exist_ok=True)
# Set up logging
logging.basicConfig(
    level=logging.DEBUG, 
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('Logs/Comparison.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger()

# Declare global variables
file_entry_previous = None
file_entry_latest = None

calculated_amount = []
# calculated_totals = 0

def load_sheets(file_previous, file_latest):
    try:
        logger.info("Loading sheets from the provided Excel files.")
        sheet_pr_previous = pd.read_excel(file_previous, sheet_name="PR")
        sheet_pr_latest = pd.read_excel(file_latest, sheet_name="PR")

        sheet_vdpmv_previous = pd.read_excel(file_previous, sheet_name="SumVDPMVReport")
        sheet_vdpmv_latest = pd.read_excel(file_latest, sheet_name="SumVDPMVReport")

        sheet_partner_previous = pd.read_excel(file_previous, sheet_name="Div2PartnerList")
        sheet_partner_latest = pd.read_excel(file_latest, sheet_name="Div2PartnerList")

        sheet_deductions_previous = pd.read_excel(file_previous, sheet_name="Deduction and other Revenue")
        sheet_deductions_latest = pd.read_excel(file_latest, sheet_name="Deduction and other Revenue")
        
        logger.info("Sheets loaded successfully.")
        return sheet_pr_previous, sheet_pr_latest, sheet_vdpmv_previous, sheet_vdpmv_latest, sheet_partner_previous, sheet_partner_latest, sheet_deductions_previous, sheet_deductions_latest
    except Exception as e:
        logger.error(f"Error loading sheets: {e}")
        raise

def clean_currency(value):
    try:
        if isinstance(value, str):
            value = value.replace('$', '').replace(',', '').strip()
            return round(float(value), 2) if value else None
        return value
    except ValueError:
        logger.error(f"Error cleaning currency value: {value}")
        return None

def calculate_totals(deductions_sheet, pr_sheet):
    try:
        calculated_totals = 0
        clients = deductions_sheet["TYPE"].unique()
        print("Clients in deductions_sheet:", clients)
        for client in clients: 
            # logger.info(f"Calculating totals for {client} clients.")

            client_header_row = pr_sheet[pr_sheet.iloc[:, 0].astype(str).str.contains(client, na=False, case=False)]
            if client_header_row.empty:
                logger.warning(f"No header row found for client {client} in pr_sheet.")
                continue  # Skip to the next client
            # Get the index of the client header row
            client_header_index = client_header_row.index[0]
            # logger.info(f"Found client header for {client} at row {client_header_index}.")

            # Find the first empty row after the client header row to determine the endpoint
            empty_row_index = pr_sheet.iloc[client_header_index + 1:, 0].isna().idxmax()

            if pd.isna(empty_row_index):
                # If no empty row is found, use the last row of the DataFrame
                next_header_index = pr_sheet.shape[0]
            else:
                # The index of the first empty row marks the endpoint
                next_header_index = empty_row_index + client_header_index + 1

            # logger.info(f"Partner rows for client {client} are between rows {client_header_index + 1} and {next_header_index - 1}.")

            # Extract the partner rows between the client header and the first empty row
            partner_rows = pr_sheet.iloc[client_header_index + 1:next_header_index]

            # Get all unique partners for this client from the deductions sheet
            total_partners = deductions_sheet["PARTNER"].unique()
            # logger.info(f"Total partners for {client}: {len(total_partners)}")

            total_amount = 0

            # Calculate total amount for matching partners
            for partner in total_partners:
                # Find rows in partner_rows matching the partner
                partner_rows_matched = partner_rows[partner_rows.iloc[:, 0].astype(str).str.strip() == str(partner).strip()]

                if not partner_rows_matched.empty:
                    # Add the amount found in column 16 (assuming it's the amount column)
                    total_amount += partner_rows_matched.iloc[0, 16]  # Column 16 contains the amount
                    # logger.info(f"Amount for partner {partner}: {partner_rows_matched.iloc[0, 14]}")
            calculated_totals += total_amount
        logger.info(f"Total amount for client {client}: {total_amount}")
        # logger.info("Calculating totals for trips, hours, operators, and amounts.")
        logger.info(f"Total calculated mamount:{calculated_totals}")
        return calculated_totals
    except Exception as e:
        logger.error(f"Error calculating totals: {e}")
        raise

def compare_totals(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing Totals between previous and latest values.")

        # Ensure inputs are numeric
        if not isinstance(sheet_previous, (int, float)) or not isinstance(sheet_latest, (int, float)):
            raise TypeError("Both sheet_previous and sheet_latest must be numeric values.")

        # Create a DataFrame to store results
        deductions_comparison = pd.DataFrame({
            "PREVIOUS": [sheet_previous],
            "LATEST": [sheet_latest],
            "CHANGE": [sheet_latest - sheet_previous]
        })

        logger.info("Totals comparison completed.")
        return deductions_comparison

    except Exception as e:
        logger.error(f"Error comparing totals: {e}")
        raise

def compare_htotalrev(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing Hourly Total Revs between previous and latest sheets.")
        
        # Select the relevant columns directly
        previous_values = sheet_previous[["PARTNER NAME", "Total Rev"]]
        latest_values = sheet_latest[["PARTNER NAME", "Total Rev"]]

        # Merge both dataframes on "PARTNER" and handle missing values with 0
        comparison = previous_values.merge(latest_values, on="PARTNER NAME", how="outer", suffixes=("_PREVIOUS", "_LATEST")).fillna(0)

        # Calculate the change in the "Total Rev"
        comparison["CHANGE"] = comparison["Total Rev_LATEST"] - comparison["Total Rev_PREVIOUS"]

        # Prepare the final dataframe for comparison
        deductions_comparison = comparison[["PARTNER NAME", "Total Rev_PREVIOUS", "Total Rev_LATEST", "CHANGE"]].drop_duplicates(subset="PARTNER NAME")

        # Rename columns
        deductions_comparison.columns = ["PARTNER", "PREVIOUS", "LATEST", "CHANGE"]

        logger.info("HTOTALREV comparison completed.")
        return deductions_comparison

    except Exception as e:
        logger.error(f"Error comparing HTOTALREV: {e}")
        raise

def compare_liftlease(sheet_previous, sheet_latest, htotalrev_df):
    try:
        logger.info("Comparing Lift Lease between previous and latest sheets.")
        
        # Group by PARTNER and sum the LIFT LEASE TOTAL
        previous_values = sheet_previous.groupby("PARTNER", as_index=False)["LIFT LEASE TOTAL"].sum()
        latest_values = sheet_latest.groupby("PARTNER", as_index=False)["LIFT LEASE TOTAL"].sum()


        # Merge both dataframes on "PARTNER" and handle missing values with 0
        comparison = previous_values.merge(latest_values, on="PARTNER", how="outer", suffixes=("_PREVIOUS", "_LATEST")).fillna(0)
        comparison = comparison.merge(htotalrev_df, on="PARTNER", how="inner").fillna(0)

        # Calculate the change in the "LIFT LEASE TOTAL"
        comparison["CHANGE"] = comparison["LIFT LEASE TOTAL_LATEST"] - comparison["LIFT LEASE TOTAL_PREVIOUS"]

        # Prepare the final dataframe for comparison
        deductions_comparison = comparison[["PARTNER", "LIFT LEASE TOTAL_PREVIOUS", "LIFT LEASE TOTAL_LATEST", "CHANGE"]].drop_duplicates(subset="PARTNER")

        # Rename columns
        deductions_comparison.columns = ["PARTNER", "PREVIOUS", "LATEST", "CHANGE"]

        logger.info("Lift Lease comparison completed.")
        return deductions_comparison

    except Exception as e:
        logger.error(f"Error comparing Lift Lease: {e}")
        raise

def compare_violations(sheet_previous, sheet_latest, htotalrev_df):
    try:
        logger.info("Comparing Lift Lease between previous and latest sheets.")
        
        # Group by PARTNER and sum the Violation
        previous_values = sheet_previous.groupby("PARTNER", as_index=False)["Violation"].sum()
        latest_values = sheet_latest.groupby("PARTNER", as_index=False)["Violation"].sum()


        # Merge both dataframes on "PARTNER" and handle missing values with 0
        comparison = previous_values.merge(latest_values, on="PARTNER", how="outer", suffixes=("_PREVIOUS", "_LATEST")).fillna(0)
        comparison = comparison.merge(htotalrev_df, on="PARTNER", how="inner").fillna(0)

        # Calculate the change in the "LIFT LEASE TOTAL"
        comparison["CHANGE"] = comparison["Violation_LATEST"] - comparison["Violation_PREVIOUS"]

        # Prepare the final dataframe for comparison
        deductions_comparison = comparison[["PARTNER", "Violation_PREVIOUS", "Violation_LATEST", "CHANGE"]].drop_duplicates(subset="PARTNER")

        # Rename columns
        deductions_comparison.columns = ["PARTNER", "PREVIOUS", "LATEST", "CHANGE"]

        logger.info("Violation comparison completed.")
        return deductions_comparison

    except Exception as e:
        logger.error(f"Error comparing Violation: {e}")
        raise

def compare_operators(sheet_previous, sheet_latest):
    try:
        logger.info("Comparing operators between previous and latest sheets.")
        
        # Extract and store unique pairs of (OPERATOR NAME, PARTNER NAME)
        operators_previous = set(sheet_previous[["OPERATOR NAME", "PARTNER NAME"]].dropna().itertuples(index=False, name=None))
        operators_latest = set(sheet_latest[["OPERATOR NAME", "PARTNER NAME"]].dropna().itertuples(index=False, name=None))

        # Identify added and removed operators
        added = operators_latest - operators_previous
        removed = operators_previous - operators_latest

        # Convert to DataFrame format
        added_list = [{"Operator Name": op, "Partner": partner, "Change": "Added"} for op, partner in added]
        removed_list = [{"Operator Name": op, "Partner": partner, "Change": "Removed"} for op, partner in removed]

        # Create DataFrames
        added_df = pd.DataFrame(added_list)
        removed_df = pd.DataFrame(removed_list)

        # Ensure DataFrames always have the necessary columns
        if added_df.empty:
            added_df = pd.DataFrame(columns=["Operator Name", "Partner", "Change"])
        if removed_df.empty:
            removed_df = pd.DataFrame(columns=["Operator Name", "Partner", "Change"])

        # Combine the results
        operator_changes_df = pd.concat([added_df, removed_df], ignore_index=True)

        logger.info("Operator comparison completed.")
        return operator_changes_df

    except Exception as e:
        logger.error(f"Error comparing operators: {e}")
        raise
    
def compare_acceptance_rate(sheet_previous, sheet_latest, week):
    try:
        logger.info(f"Comparing Acceptance Rate for {week}.")

        # Filter data for the specific week
        prev_week = sheet_previous[sheet_previous["WeekN"] == week]
        latest_week = sheet_latest[sheet_latest["WeekN"] == week]

        # Group by "PARTNER NAME" and sum Acceptance Rate
        prev_values = prev_week.groupby("PARTNER NAME", as_index=False)["Acceptance Rate"].sum()
        latest_values = latest_week.groupby("PARTNER NAME", as_index=False)["Acceptance Rate"].sum()

        # Merge both datasets
        comparison = prev_values.merge(
            latest_values, on="PARTNER NAME", how="outer", suffixes=("_PREVIOUS", "_LATEST")
        ).fillna(0)

        # Calculate the change
        comparison["CHANGE"] = comparison["Acceptance Rate_LATEST"] - comparison["Acceptance Rate_PREVIOUS"]

        # Rename columns
        comparison.columns = ["PARTNER", "PREVIOUS", "LATEST", "CHANGE"]

        logger.info(f"{week} Acceptance Rate comparison completed.")
        return comparison

    except Exception as e:
        logger.error(f"Error comparing Acceptance Rate for {week}: {e}")
        raise

# def compare_acceptance_rate(sheet_previous, sheet_latest, week):
#     try:
#         logger.info(f"Comparing Acceptance Rate for {week}.")

#         # Filter data for the specific week
#         prev_week = sheet_previous[sheet_previous["WeekN"] == week]
#         latest_week = sheet_latest[sheet_latest["WeekN"] == week]

#         # Group by "PARTNER NAME" and sum Acceptance Rate
#         prev_values = prev_week.groupby("PARTNER NAME", as_index=False)["Acceptance Rate"].sum()
#         latest_values = latest_week.groupby("PARTNER NAME", as_index=False)["Acceptance Rate"].sum()

#         # Merge both datasets
#         comparison = prev_values.merge(
#             latest_values, on="PARTNER NAME", how="outer", suffixes=("_PREVIOUS", "_LATEST")
#         ).fillna(0)

#         # Calculate the change
#         comparison["CHANGE"] = comparison["Acceptance Rate_LATEST"] - comparison["Acceptance Rate_PREVIOUS"]

#         # Rename columns
#         comparison.columns = ["PARTNER", "PREVIOUS", "LATEST", "CHANGE"]

#         logger.info(f"{week} Acceptance Rate comparison completed.")
#         return comparison

#     except Exception as e:
#         logger.error(f"Error comparing Acceptance Rate for {week}: {e}")
#         raise

# def compare_cancellation_rate(sheet_previous, sheet_latest, week):
#     try:
#         logger.info(f"Comparing Cancellation Rate for {week}.")

#         # Filter data for the specific week
#         prev_week = sheet_previous[sheet_previous["WeekN"] == week]
#         latest_week = sheet_latest[sheet_latest["WeekN"] == week]

#         # Group by "PARTNER NAME" and sum Cancellation Rate
#         prev_values = prev_week.groupby("PARTNER NAME", as_index=False)["Cancellation Rate"].sum()
#         latest_values = latest_week.groupby("PARTNER NAME", as_index=False)["Cancellation Rate"].sum()

#         # Merge both datasets
#         comparison = prev_values.merge(
#             latest_values, on="PARTNER NAME", how="outer", suffixes=("_PREVIOUS", "_LATEST")
#         ).fillna(0)

#         # Calculate the change
#         comparison["CHANGE"] = comparison["Cancellation Rate_LATEST"] - comparison["Cancellation Rate_PREVIOUS"]

#         # Rename columns
#         comparison.columns = ["PARTNER", "PREVIOUS", "LATEST", "CHANGE"]

#         logger.info(f"{week} Cancellation Rate comparison completed.")
#         return comparison

#     except Exception as e:
#         logger.error(f"Error comparing Cancellation Rate for {week}: {e}")
#         raise

# def compare_utilization(sheet_previous, sheet_latest, week):
#     try:
#         logger.info(f"Comparing Utilization% for {week}.")

#         # Filter data for the specific week
#         prev_week = sheet_previous[sheet_previous["WeekN"] == week]
#         latest_week = sheet_latest[sheet_latest["WeekN"] == week]

#         # Group by "PARTNER NAME" and sum Utilization%
#         prev_values = prev_week.groupby("PARTNER NAME", as_index=False)["Utilization%"].sum()
#         latest_values = latest_week.groupby("PARTNER NAME", as_index=False)["Utilization%"].sum()

#         # Merge both datasets
#         comparison = prev_values.merge(
#             latest_values, on="PARTNER NAME", how="outer", suffixes=("_PREVIOUS", "_LATEST")
#         ).fillna(0)

#         # Calculate the change
#         comparison["CHANGE"] = comparison["Utilization%_LATEST"] - comparison["Utilization%_PREVIOUS"]

#         # Rename columns
#         comparison.columns = ["PARTNER", "PREVIOUS", "LATEST", "CHANGE"]

#         logger.info(f"{week} Utilization% comparison completed.")
#         return comparison

#     except Exception as e:
#         logger.error(f"Error comparing Utilization% for {week}: {e}")
#         raise

# def compare_week1(SumVDPMV_Sheet):
#     try:
#         logger.info("Comparing weeks between previous and latest sheets.")

#         # Filter for "Week1" only
#         sheet_week1 = SumVDPMV_Sheet[SumVDPMV_Sheet["WeekN"] == "Week1"]
        
#         # Group by week1 and sum the values
#         grouped_week1 = sheet_week1.groupby("WeekN")[["Acceptance Rate", "Cancellation Rate", "Utilization%", "Payable Normal Hours", "Payable Bonus Hours"]].sum()

#         logger.info("Week1 comparison completed.")
#         # return deductions_comparison
#         return {
#             "Acceptance Rate": grouped_week1["Acceptance Rate"].sum(),
#             "Cancellation Rate": grouped_week1["Cancellation Rate"].sum(),
#             "Utilization%": grouped_week1["Utilization%"].sum(),
#             "Payable Normal Hours": grouped_week1["Payable Normal Hours"].sum(),
#             "Payable Bonus Hours": grouped_week1["Payable Bonus Hours"].sum(),
#         }

#     except Exception as e:
#         logger.error(f"Error comparing Week1: {e}")
#         raise

# def compare_week2(SumVDPMV_Sheet):
#     try:
#         logger.info("Comparing weeks between previous and latest sheets.")

#         # Filter for "Week1" only
#         sheet_week1 = SumVDPMV_Sheet[SumVDPMV_Sheet["WeekN"] == "Week2"]
        
#         # Group by week1 and sum the values
#         grouped_week1 = sheet_week1.groupby("WeekN")[["Acceptance Rate", "Cancellation Rate", "Utilization%", "Payable Normal Hours", "Payable Bonus Hours"]].sum()

#         logger.info("Week1 comparison completed.")
#         # return deductions_comparison
#         return {
#             "Acceptance Rate": grouped_week1["Acceptance Rate"].sum(),
#             "Cancellation Rate": grouped_week1["Cancellation Rate"].sum(),
#             "Utilization%": grouped_week1["Utilization%"].sum(),
#             "Payable Normal Hours": grouped_week1["Payable Normal Hours"].sum(),
#             "Payable Bonus Hours": grouped_week1["Payable Bonus Hours"].sum(),
#         }

#     except Exception as e:
#         logger.error(f"Error comparing Week1: {e}")
#         raise

def apply_formatting(sheet_name, wb):
    try:
        logger.info(f"Applying formatting to sheet: {sheet_name}.")
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                if ws[1][cell.column - 1].value.lower() == "change":
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100")
                        elif cell.value < 0:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006")
                    elif isinstance(cell.value, str):
                        if cell.value.lower() in ["increased", "added"]:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(color="006100")
                        elif cell.value.lower() in ["decreased", "removed"]:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(color="9C0006")
        logger.info(f"Formatting applied successfully to sheet: {sheet_name}.")
    except Exception as e:
        logger.error(f"Error applying formatting to sheet {sheet_name}: {e}")
        raise


def save_comparison_results(output_folder, comparison_data, filename):
    try:
        logger.info(f"Saving comparison results to {filename}.")
        os.makedirs(output_folder, exist_ok=True)
        full_comparison_file = os.path.join(output_folder, filename)
        with pd.ExcelWriter(full_comparison_file, engine="openpyxl") as writer:
            for sheet_name, data in comparison_data.items():
                data.to_excel(writer, sheet_name=sheet_name, index=False)

        wb_full = load_workbook(full_comparison_file)
        for sheet in comparison_data.keys():
            apply_formatting(sheet, wb_full)
        wb_full.save(full_comparison_file)
        wb_full.close()

        logger.info(f"Comparison results saved successfully to {filename}.")
    except Exception as e:
        logger.error(f"Error saving comparison results: {e}")
        raise


def main(file_previous, file_latest):
    try: 
        file_entry_previous = file_previous
        file_entry_latest = file_latest

        print(f"{file_entry_previous} + {file_entry_latest}")
        logger.info("Starting main comparison process.")
        output_folder = "ComparedResults"
        os.makedirs(output_folder, exist_ok=True)

        sheet_pr_previous, sheet_pr_latest, sheet_vdpmv_previous, sheet_vdpmv_latest, sheet_partner_previous, sheet_partner_latest, sheet_deductions_previous, sheet_deductions_latest = load_sheets(file_previous, file_latest)

        # 1. Process the data without any filtering (full comparison)

        # Compare totals
        prev_totals = calculate_totals(sheet_deductions_previous, sheet_pr_previous)
        lat_totals = calculate_totals(sheet_deductions_latest, sheet_pr_latest)
        totals_comparison_df = compare_totals(prev_totals, lat_totals)
        # logger.info(f"Totals DF: {totals_comparison_df}")

        # 2. Compare hourly total revs
        compare_htotalrev_df = compare_htotalrev(sheet_vdpmv_previous, sheet_vdpmv_latest)
        # logger.info(f"HTOTAL_REV DF: {compare_htotalrev_df}")

        # 3. Compare Lift Lease
        compare_liftlease_df = compare_liftlease(sheet_deductions_previous, sheet_deductions_latest, compare_htotalrev_df)
        # logger.info(f"Lift Lease DF: {compare_liftlease_df}")

        # 4. Compare Violations
        compare_violations_df = compare_violations(sheet_deductions_previous, sheet_deductions_latest, compare_htotalrev_df)
        # logger.info(f"Violations DF: {compare_violations_df}")

        # 5. Compare operators
        operator_changes_df = compare_operators(sheet_partner_previous, sheet_partner_latest)
        # logger.info(f"Operator Changes DF:{operator_changes_df}")

        # 6. Week 1 comparison
        # week1_comparison_previous = compare_week1(sheet_vdpmv_previous)
        # week1_comparison_latest = compare_week1(sheet_vdpmv_latest)

        # # Calculate differences and changes
        # differences = {
        #     "Acceptance Rate": week1_comparison_latest["Acceptance Rate"] - week1_comparison_previous["Acceptance Rate"],
        #     "Cancellation Rate": week1_comparison_latest["Cancellation Rate"] - week1_comparison_previous["Cancellation Rate"],
        #     "Utilization%": week1_comparison_latest["Utilization%"] - week1_comparison_previous["Utilization%"],
        #     "Payable Normal Hours": week1_comparison_latest["Payable Normal Hours"] - week1_comparison_previous["Payable Normal Hours"],
        #     "Payable Bonus Hours": week1_comparison_latest["Payable Bonus Hours"] - week1_comparison_previous["Payable Bonus Hours"],
        # }
        # changes = {key: "Increased" if diff > 0 else "Decreased" if diff < 0 else "No Change"
        #             for key, diff in differences.items()}

        # # Create summary DataFrame
        # summary_table1 = {
        #     "Metric": ["Acceptance Rate", "Cancellation Rate", "Utilization%", "Payable Normal Hours", "Payable Bonus Hours"],
        #     "Previous": [week1_comparison_previous[key] for key in week1_comparison_previous],
        #     "Latest": [week1_comparison_latest[key] for key in week1_comparison_latest ],
        #     "Difference": [differences[key] for key in differences],
        #     "Change": [changes[key] for key in changes],
        # }
        # week1_summary_df = pd.DataFrame(summary_table1)
        # logger.info(f"Summary DF: {week1_summary_df}")

        # # 6. Week 2 comparison
        # week2_comparison_previous = compare_week2(sheet_vdpmv_previous)
        # week2_comparison_latest = compare_week2(sheet_vdpmv_latest)

        # # Calculate differences and changes
        # differences = {
        #     "Acceptance Rate": week2_comparison_latest["Acceptance Rate"] - week2_comparison_previous["Acceptance Rate"],
        #     "Cancellation Rate": week2_comparison_latest["Cancellation Rate"] - week2_comparison_previous["Cancellation Rate"],
        #     "Utilization%": week2_comparison_latest["Utilization%"] - week2_comparison_previous["Utilization%"],
        #     "Payable Normal Hours": week2_comparison_latest["Payable Normal Hours"] - week2_comparison_previous["Payable Normal Hours"],
        #     "Payable Bonus Hours": week2_comparison_latest["Payable Bonus Hours"] - week2_comparison_previous["Payable Bonus Hours"],
        # }
        # changes = {key: "Increased" if diff > 0 else "Decreased" if diff < 0 else "No Change"
        #             for key, diff in differences.items()}

        # # Create summary DataFrame
        # summary_table2 = {
        #     "Metric": ["Acceptance Rate", "Cancellation Rate", "Utilization%", "Payable Normal Hours", "Payable Bonus Hours"],
        #     "Previous": [week2_comparison_previous[key] for key in week2_comparison_previous],
        #     "Latest": [week2_comparison_latest[key] for key in week2_comparison_latest ],
        #     "Difference": [differences[key] for key in differences],
        #     "Change": [changes[key] for key in changes],
        # }
        # week2_summary_df = pd.DataFrame(summary_table2)
        # logger.info(f"Summary DF: {week2_summary_df}")

        # Save the full comparison results
        full_comparison_file = os.path.join(output_folder, "DIV2_Tables.xlsx")
        excel_sheets = []
        with pd.ExcelWriter(full_comparison_file, engine="openpyxl") as writer:
            totals_comparison_df.to_excel(writer, sheet_name="TotalInvoicePayment", index=False)
            excel_sheets.append("TotalInvoicePayment")
            compare_htotalrev_df.to_excel(writer, sheet_name="HTotalRevComparison", index=False)
            excel_sheets.append("HTotalRevComparison")
            compare_liftlease_df.to_excel(writer, sheet_name="LiftLeaseComparison", index=False)
            excel_sheets.append("LiftLeaseComparison")
            compare_violations_df.to_excel(writer, sheet_name="ViolationComparison", index=False)
            excel_sheets.append("ViolationComparison")
            operator_changes_df.to_excel(writer, sheet_name="OperatorChanges", index=False)
            excel_sheets.append("OperatorChanges")
            for week in ["Week1", "Week2"]:
                acceptance_changes_df = compare_acceptance_rate(sheet_vdpmv_previous, sheet_vdpmv_latest, week)
                acceptance_changes_df.to_excel(writer, sheet_name="{week}AcceptanceRateComparison", index=False)
                excel_sheets.append(f"{week}AcceptanceRateComparison")
            # week1_summary_df.to_excel(writer, sheet_name="Week1Comparison", index=False)
            # week2_summary_df.to_excel(writer, sheet_name="Week2Comparison", index=False)

        # Doperator_changes_df.to_excel(writer, sheet_name="DateComparison", index=False)

        # Apply formatting to the full comparison file
        wb_full = load_workbook(full_comparison_file)
        for sheet in excel_sheets:                                                                                                
            apply_formatting(sheet, wb_full)
        wb_full.save(full_comparison_file)
        wb_full.close()

        logger.info(f"Main comparison process completed successfully. File saved to {full_comparison_file}.")
        # time.sleep(2)
        # db.main(file_previous, file_latest)
    except Exception as e:
        logger.error(f"Error in main comparison process: {e}")
        raise
    # finally:
    #     wb_client.close()


def open_file_dialog(entry):
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsm;*.xlsx")])
    if filename:
        entry.delete(0, tk.END)
        entry.insert(0, filename)


def create_gui():
    global file_entry_previous, file_entry_latest # Declare them as global variables

    # Set up the GUI window
    root = tk.Tk()
    root.title("Comparison Tool")

    # Create and place labels, entry boxes, and buttons
    tk.Label(root, text="Previous File:").grid(row=0, column=0, padx=10, pady=5)
    entry_previous = tk.Entry(root, width=50)
    entry_previous.grid(row=0, column=1, padx=10, pady=5)
    
    # tk.Button(root, text="Browse", command=lambda: open_file_dialog(entry_previous)).grid(row=0, column=2, padx=10, pady=5)

    tk.Label(root, text="Latest File:").grid(row=1, column=0, padx=10, pady=5)
    entry_latest = tk.Entry(root, width=50)
    entry_latest.grid(row=1, column=1, padx=10, pady=5)
    
    # tk.Button(root, text="Browse", command=lambda: open_file_dialog(entry_latest)).grid(row=1, column=2, padx=10, pady=5)

    # Button to trigger the comparison process
    # tk.Button(root, text="Compare", command=lambda: (main(entry_previous.get(), entry_latest.get()), root.destroy())).grid(row=2, column=1, pady=20)
    tk.Button(root, text="Compare", command=lambda: handle_comparison(entry_previous.get(), entry_latest.get(), root)).grid(row=2, column=1, pady=20)

    def handle_comparison(file_previous, file_latest, root):
        try:
            main(file_previous, file_latest)
        except Exception as e:
            print(f"An error occurred: {e}")
            # Check for the disconnection error and close the GUI if it happens
            if isinstance(e, OSError) and "The object invoked has disconnected" in str(e):
                print("Disconnected from Excel, closing GUI.")
                root.quit()  # This will close the Tkinter window
        finally:
            root.destroy()  # Close the window in all cases
    # Start the GUI loop
    root.mainloop()


if __name__ == "__main__":
    create_gui()

