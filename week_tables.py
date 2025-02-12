import xlwings as xw
import openpyxl
from xlwings.utils import rgb_to_int
import win32com.client
from openpyxl.drawing.image import Image
from PIL import Image as PILImage, ImageDraw, ImageFont
import os
import logging
import time
import sys

# Set up logging
log_folder = "Logs"
os.makedirs(log_folder, exist_ok=True)
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(log_folder, 'Comparison.log')),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger()

def average(column_values):
    values = [cell.value for row in column_values for cell in row if cell.value is not None]
    return round(sum(values) / len(values), 2) if values else 0.00  # Returns a float

# Define the comparison files and corresponding sheets
def main(file_previous, file_latest): 
    logger.info(f"Week {file_previous} + {file_latest}")
    comparison_files = [
        ('ComparedResults/DIV2_Tables.xlsx', 'Week1'),
        ('ComparedResults/DIV2_Tables.xlsx', 'Week2')
    ]

    # Open the Dashboard workbook
    dashboard_file = 'ComparedResults/Dashboard.xlsm'
    app = xw.App(visible=True)

    try:
        # Open the existing workbook (Dashboard)
        wb_dashboard = app.books.open(dashboard_file)

        # Loop through each comparison file and corresponding sheet
        for comparison_file, sheet_name in comparison_files:
            logger.info(f"Processing {comparison_file} -> {sheet_name}")

            # Load the comparison workbook and sheet for week 1
            wb_comparison = openpyxl.load_workbook(comparison_file)
            sheet_Week1AcceptRateComp = wb_comparison[f'{sheet_name}AcceptRateComp']
            sheet_Week1CancelRateComp = wb_comparison[f'{sheet_name}CancelRateComp']
            sheet_Week1UtilizationComp = wb_comparison[f'{sheet_name}UtilizationComp']
            sheet_Week1PNormalHrsComp = wb_comparison[f'{sheet_name}PNormalHrsComp']
            sheet_Week1PBonusHrsComp = wb_comparison[f'{sheet_name}PBonusHrsComp']
            sheet_Week1ReqHrsComp = wb_comparison[f'{sheet_name}ReqHrsComp']
            logger.info(f"Processing {comparison_file} -> {sheet_name}")

            # # Retrieve values from the 'Week1AcceptRateComp' sheet in the comparison file
            # prev_Accept = f"{sum(cell.value for row in sheet_Week1AcceptRateComp.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):.2f}"
            # lat_Accept = f"{sum(cell.value for row in sheet_Week1AcceptRateComp.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):.2f}"
            # diff_Accept = f"{sum(cell.value for row in sheet_Week1AcceptRateComp.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):.2f}"

            # # Retrieve values from the 'Week1CancelRateComp' sheet in the comparison file
            # prev_Cancel = f"{sum(cell.value for row in sheet_Week1CancelRateComp.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):.2f}"
            # lat_Cancel= f"{sum(cell.value for row in sheet_Week1CancelRateComp.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):.2f}"
            # diff_Cancel = f"{sum(cell.value for row in sheet_Week1CancelRateComp.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):.2f}"

            # # Retrieve values from the 'Week1UtilizationComp' sheet in the comparison file
            # prev_UtilizationComp = f"{sum(cell.value for row in sheet_Week1UtilizationComp.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None):.2f}"
            # lat_UtilizationComp = f"{sum(cell.value for row in sheet_Week1UtilizationComp.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None):.2f}"
            # diff_UtilizationComp = f"{sum(cell.value for row in sheet_Week1UtilizationComp.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None):.2f}"

            # Retrieve values from the 'Week1AcceptRateComp' sheet in the comparison file
            prev_Accept = f"{(sum(values) / len(values)):.2f}" if (values := [cell.value for row in sheet_Week1AcceptRateComp.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None]) else "0.00"
            lat_Accept = f"{(sum(values) / len(values)):.2f}" if (values := [cell.value for row in sheet_Week1AcceptRateComp.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None]) else "0.00"
            diff_Accept = f"{(sum(values) / len(values)):.2f}" if (values := [cell.value for row in sheet_Week1AcceptRateComp.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None]) else "0.00"

            # Retrieve values from the 'Week1CancelRateComp' sheet in the comparison file
            prev_Cancel = f"{(sum(values) / len(values)):.2f}" if (values := [cell.value for row in sheet_Week1CancelRateComp.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None]) else "0.00"
            lat_Cancel = f"{(sum(values) / len(values)):.2f}" if (values := [cell.value for row in sheet_Week1CancelRateComp.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None]) else "0.00"
            diff_Cancel = f"{(sum(values) / len(values)):.2f}" if (values := [cell.value for row in sheet_Week1CancelRateComp.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None]) else "0.00"

            # Retrieve values from the 'Week1UtilizationComp' sheet in the comparison file
            prev_UtilizationComp = f"{(sum(values) / len(values)):.2f}" if (values := [cell.value for row in sheet_Week1UtilizationComp.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None]) else "0.00"
            lat_UtilizationComp = f"{(sum(values) / len(values)):.2f}" if (values := [cell.value for row in sheet_Week1UtilizationComp.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None]) else "0.00"
            diff_UtilizationComp = f"{(sum(values) / len(values)):.2f}" if (values := [cell.value for row in sheet_Week1UtilizationComp.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None]) else "0.00"


            # Retrieve values from the 'Week1PNormalHrsComp' sheet in the comparison file
            prev_Week1PNormalHrsComp = f"{sum(cell.value for row in sheet_Week1PNormalHrsComp.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None) or 0:,.2f}"
            lat_Week1PNormalHrsComp = f"{sum(cell.value for row in sheet_Week1PNormalHrsComp.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None) or 0:,.2f}"
            diff_Week1PNormalHrsComp = f"{sum(cell.value for row in sheet_Week1PNormalHrsComp.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None) or 0:,.2f}"

            # Retrieve values from the 'Week1PBonusHrsComp' sheet in the comparison file
            prev_Week1PBonusHrsComp = f"{sum(cell.value for row in sheet_Week1PBonusHrsComp.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None) or 0:,.2f}"
            lat_Week1PBonusHrsComp = f"{sum(cell.value for row in sheet_Week1PBonusHrsComp.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None) or 0:,.2f}"
            diff_Week1PBonusHrsComp = f"{sum(cell.value for row in sheet_Week1PBonusHrsComp.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None) or 0:,.2f}"

            # Retrieve values from the 'Week1ReqHrsComp' sheet in the comparison file
            prev_Week1ReqHrsComp = f"{(sum(values) / len(values)):.2f}" if (values := [cell.value for row in sheet_Week1ReqHrsComp.iter_rows(min_row=2, max_row=50, min_col=2, max_col=2) for cell in row if cell.value is not None]) else "0.00"
            lat_Week1ReqHrsComp = f"{(sum(values) / len(values)):.2f}" if (values := [cell.value for row in sheet_Week1ReqHrsComp.iter_rows(min_row=2, max_row=50, min_col=3, max_col=3) for cell in row if cell.value is not None]) else "0.00"
            diff_Week1ReqHrsComp = f"{(sum(values) / len(values)):.2f}" if (values := [cell.value for row in sheet_Week1ReqHrsComp.iter_rows(min_row=2, max_row=50, min_col=4, max_col=4) for cell in row if cell.value is not None]) else "0.00"

            logger.info(f"Accept Hrs: {prev_Accept} -> {lat_Accept}")

            # Get the corresponding sheet in the dashboard
            sheet_dashboard = wb_dashboard.sheets[sheet_name]
            sheet_dashboard.activate()

            # Ensure the dashboard sheet is active
            # wb_dashboard.app.api.ActiveSheet = sheet_dashboard.api 
            # wb_dashboard.app.api.Application.ScreenUpdating = True 

            # Access the Week1AcceptRateComp shape via the API and set the value
            txt_Accept = sheet_dashboard.shapes['txtDAcceptRateDiff'].api
            txt_Accept.TextFrame2.TextRange.Text = f"{prev_Accept}% to {lat_Accept}%"
            txt_Accept_diff = sheet_dashboard.shapes['txtAcceptDiff'].api
            txt_Accept_diff.TextFrame2.TextRange.Text = f"{diff_Accept}%"

            # Access the Week1CancelRateComp shape via the API and set the value
            txt_Cancel = sheet_dashboard.shapes['txtDCancelDiff'].api
            txt_Cancel.TextFrame2.TextRange.Text = f"{prev_Cancel}% to {lat_Cancel}%"
            txt_Cancel_diff = sheet_dashboard.shapes['txtCancelDiff'].api
            txt_Cancel_diff.TextFrame2.TextRange.Text = f"{diff_Cancel}%"

            # Access the Week1UtilizationComp shape via the API and set the value
            txt_UtilizationComp = sheet_dashboard.shapes['txtDUtilizationDiff'].api
            txt_UtilizationComp.TextFrame2.TextRange.Text = f"{prev_UtilizationComp}% to {lat_UtilizationComp}%"
            txt_UtilizationComp_diff = sheet_dashboard.shapes['txtUtilizationDiff'].api
            txt_UtilizationComp_diff.TextFrame2.TextRange.Text = f"{diff_UtilizationComp}%"

            # Access the Week1PNormalHrsComp shape via the API and set the value
            txt_PNormal = sheet_dashboard.shapes['txtDPNormalHrsDiff'].api
            txt_PNormal.TextFrame2.TextRange.Text = f"${prev_Week1PNormalHrsComp} to ${lat_Week1PNormalHrsComp}"
            txt_PNormal_diff = sheet_dashboard.shapes['txtPNormalHrsDiff'].api
            txt_PNormal_diff.TextFrame2.TextRange.Text = f"${diff_Week1PNormalHrsComp}"

            # Access the Week1PBonusHrsComp shape via the API and set the value
            txt_PBonus = sheet_dashboard.shapes['txtDPBonuslHrsDiff'].api
            txt_PBonus.TextFrame2.TextRange.Text = f"${prev_Week1PBonusHrsComp} to ${lat_Week1PBonusHrsComp}"
            txt_PBonus_diff = sheet_dashboard.shapes['txtPBonusHrsDiff'].api
            txt_PBonus_diff.TextFrame2.TextRange.Text = f"${diff_Week1PBonusHrsComp}"

            # Access the Week1ReqHrsComp shape via the API and set the value
            txt_ReqHrs = sheet_dashboard.shapes['txtDReqHrsDiff'].api
            txt_ReqHrs.TextFrame2.TextRange.Text = f"{prev_Week1ReqHrsComp}% to {lat_Week1ReqHrsComp}%"
            txt_ReqHrs_diff = sheet_dashboard.shapes['txtReqHrsDiff'].api
            txt_ReqHrs_diff.TextFrame2.TextRange.Text = f"{diff_Week1ReqHrsComp}%"


            # Run the VBA macro to update the color based on the values
            try:
                # Parameters: TextBox names and corresponding values
                textBoxNames = ["txtAcceptDiff", "txtCancelDiff", "txtUtilizationDiff", "txtPNormalHrsDiff", "txtPBonusHrsDiff", "txtReqHrsDiff"]
                values = [diff_Accept, diff_Cancel, diff_UtilizationComp, diff_Week1PNormalHrsComp, diff_Week1PBonusHrsComp, diff_Week1ReqHrsComp]

                # Loop through the text boxes and update colors based on the values
                for i, textBoxName in enumerate(textBoxNames):
                    wb_dashboard.macro("UpdateSummaryColor")(sheet_name, textBoxName, values[i])
                    logger.info(f"Successfully updated color for {textBoxName} with value '{values[i]}'.")
            except Exception as e:
                logger.error(f"A Week TextBox error occurred: {e}")
        # wb_comparison.save()
        # wb_comparison.close()
        # app.api.ScreenUpdating = True  # Enable screen updating
        # Save the changes to the dashboard workbook
        # wb_dashboard.refresh_all() # Refresh all data connections
        time.sleep(2)
        wb_dashboard.save()
        wb_dashboard.close()
        logger.info(f"{dashboard_file} has been successfully updated and saved.")

        # # paste_picture(comparison_files, dashboard_file)
        app.quit()
        time.sleep(2)
        wpaste_picture()

    except Exception as e:
        logger.info(f"An Week error occurred: {e}")
    # finally:
    #     wb_dashboard.save()
    #     wb_dashboard.close()
        # app.quit()
        
        # Reopen the Excel file
        # app = xw.App(visible=True)  # Open Excel with the app visible
        # wb_dashboard = app.books.open(dashboard_file)  # Reopen the file


def wpaste_picture():
    comparison_files = [
        ('ComparedResults\\DIV2_Tables.xlsx', 'Week1'),
        ('ComparedResults\\DIV2_Tables.xlsx', 'Week2')
    ]
    
    relative_dashboard_path = "ComparedResults\\Dashboard.xlsm"
    # Get the absolute path of the current script's directory
    # script_dir = os.path.dirname(os.path.realpath(__file__))

    # # Build the full path to the dashboard file by joining the script directory and the relative path
    # dashboard_file = os.path.join(script_dir, relative_dashboard_path)

    # Get the base directory correctly whether running as script or PyInstaller executable
    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)  # Executable folder
    else:
        script_dir = os.path.dirname(os.path.realpath(__file__))  # Script folder

    dashboard_file = os.path.join(script_dir, "ComparedResults", "Dashboard.xlsm")

    excel = None

    try:
        # Initialize Excel application
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True  # Set to True for debugging

        # Check if the file exists
        if not os.path.exists(dashboard_file):
            logger.info(f"Error: Dashboard file does not exist at {dashboard_file}")
            return
        
        # Open the Dashboard workbook
        wb_dashboard = excel.Workbooks.Open(dashboard_file)
        if wb_dashboard is None:
            logger.info(f"Failed to open the Dashboard workbook at {dashboard_file}")
            return

        # Delete existing pictures if they exist
        for target_sheet_name in ['Week1','Week2']:
            ws_dashboard = wb_dashboard.Sheets(target_sheet_name)
            ws_dashboard.Activate()
            for picture_name in ['AcceptTable', 'CancelTable', 'UtilizationTable', 'NormalTable', 'BonusTable', 'ReqTable']:
                try:
                    ws_dashboard.Shapes(picture_name).Delete()  # Attempt to delete the picture
                    logger.info(f"Deleted existing picture: {picture_name} in {target_sheet_name}")
                except Exception:
                    logger.info(f"No existing picture named {picture_name} found in {target_sheet_name}")  

        time.sleep(2)
        
        # Process each comparison file
        for comparison_file, target_sheet_name in comparison_files:
            # Target cells for each sheet in the comparison file
            target_cells = {
                f'{target_sheet_name}AcceptRateComp': (9, 13),
                f'{target_sheet_name}CancelRateComp': (9, 21),
                f'{target_sheet_name}UtilizationComp': (9, 29),
                f'{target_sheet_name}PNormalHrsComp': (28, 13),
                f'{target_sheet_name}PBonusHrsComp': (28, 21),
                f'{target_sheet_name}ReqHrsComp': (28, 29)
            }
            # Build the full path for the comparison file
            comparison_file_path = os.path.join(script_dir, comparison_file)
            
            # Check if the comparison file exists
            if not os.path.exists(comparison_file_path):
                logger.info(f"Error: Comparison file does not exist at {comparison_file_path}")
                continue

            # Open the comparison workbook
            wb_comparison = excel.Workbooks.Open(comparison_file_path)
            if wb_comparison is None:
                logger.info(f"Failed to open the comparison workbook at {comparison_file_path}")
                continue

            # Process each sheet in the comparison file (Week1UtilizationComp, Week1AcceptRateComp, Week1PNormalHrsComp)
            for sheet_name, target_cell in target_cells.items():
                sheet = wb_comparison.Sheets(sheet_name)
                table_width = 0
                table_height = 0
                table_name = None
                if sheet is None:
                    logger.info(f"Failed to access the '{sheet_name}' sheet in {comparison_file_path}")
                    continue

                # Get the used range
                used_range = sheet.UsedRange

                time.sleep(1)
                if used_range.Rows.Count > 1 and used_range.Columns.Count > 1:
                    try:
                        table_width = (used_range.Width) # Get the width of the used range
                        table_height = (used_range.Height) # Get the height of the used range
                        used_range.CopyPicture(Format=2)
                        logger.info(f"Copied picture from {sheet_name}, width: {table_width*0.0352778:.2f} cm, height: {table_height*0.0352778:.2f} cm")
                    except Exception as e:
                        logger.error(f"Failed to copy picture from {sheet_name}: {e}")
                        continue
                else:
                    logger.error(f"Skipping {sheet_name}: No data in the range.")
                    continue

                # Activate the target sheet in the Dashboard workbook
                ws_dashboard = wb_dashboard.Sheets(target_sheet_name)
                if ws_dashboard is None:
                    logger.info(f"Failed to access the sheet '{target_sheet_name}' in the Dashboard workbook.")
                    wb_comparison.Close(SaveChanges=False)
                    continue
                
                time.sleep(1)
                # Paste as a picture into the target sheet
                ws_dashboard.Activate()
                row, col = target_cell
                target_cell_range = ws_dashboard.Cells(row, col)  # Adjust as needed
                ws_dashboard.Paste(target_cell_range)

                # Position and resize the pasted picture
                pasted_picture = ws_dashboard.Shapes(ws_dashboard.Shapes.Count)
                pasted_picture.Left = target_cell_range.Left
                pasted_picture.Top = target_cell_range.Top
                
                logger
                # Name the pasted picture according to the sheet
                if sheet_name == f'{target_sheet_name}UtilizationComp':
                    pasted_picture.Name = 'UtilizationTable'
                elif sheet_name == f'{target_sheet_name}AcceptRateComp':
                    pasted_picture.Name = 'AcceptTable'
                elif sheet_name == f'{target_sheet_name}PNormalHrsComp':
                    pasted_picture.Name = 'NormalTable'
                elif sheet_name == f'{target_sheet_name}CancelRateComp':
                    pasted_picture.Name = 'CancelTable'
                elif sheet_name == f'{target_sheet_name}PBonusHrsComp':
                    pasted_picture.Name = 'BonusTable'
                elif sheet_name == f'{target_sheet_name}ReqHrsComp':
                    pasted_picture.Name = 'ReqTable'

                table_name = pasted_picture.Name
                logger.info(f"Table Name: {table_name}")

                time.sleep(1)
                logger.info(f"'{pasted_picture.Name}' successfully pasted in the {sheet_name} Sheet.")

                # Adjust the container size based on the table size
                # Adjust only the container that matches the table name
                if table_name:
                    container_name = table_name.replace("Table", "Container")  # Match the container name
                    try:
                        container = ws_dashboard.Shapes(container_name)
                        # container.Width = table_width + 95  # Add 3.35 cm to width
                        container.Width = table_width - 1  # Add 3.35 cm to width
                        # container.Height = table_height + 123  # Add 4.33 cm to height
                        container.Height = table_height + 56  # Add 4.33 cm to height
                        logger.info(f"Resized {container_name} to width: {(container.Width)*0.0352778:.2f} cm, height: {(container.Height)*0.0352778:.2f} cm")
                    except Exception as e:
                        logger.error(f"Failed to resize {container_name}: {e}")

                # wb_comparison.Save()
                # wb_dashboard.Save()

        wb_comparison.Save()
        wb_comparison.Close() 
            # Close the comparison workbook without saving
            # wb_comparison.Close(SaveChanges=True)

        # Save and close the Dashboard workbook
        wb_dashboard.Save()
        wb_dashboard.Close()
        excel.Quit()

        logger.info("Data pasted as pictures successfully.")

        # app = xw.App(visible=True)  # Open Excel with the app visible
        # wb_dashboard = app.books.open(dashboard_file)  # Reopen the file
    except Exception as e:
        logger.error(f"An error occurred: {e}")

        if 'wb_dashboard' in locals() and wb_dashboard:
            wb_dashboard.Close(SaveChanges=False)

        if excel:
            excel.Quit()
            del excel
        # logger.error(f"An error occurred: {e}")
        # wb_dashboard.Close()
        # excel.Quit()
        # del excel 
    finally:
        # Ensure Excel is properly quit and the object is released
        if excel:
            excel.Quit()  # Close the Excel application
            del excel 
        app = xw.App(visible=True)  # Open Excel with the app visible
        wb_dashboard = app.books.open(dashboard_file)  # Reopen the file

# if __name__ == '__main__':
    
#     comparison_files = [
#         ('Compared Results/Full_Comparison.xlsx', 'Dashboard'),
#         ('Compared Results/CCCTA_Comparison.xlsx', 'CCCTA'),
#         ('Compared Results/LAVTA_Comparison.xlsx', 'LAVTA')
#     ]
#     dashboard_file = 'Compared Results/Dashboard.xlsm'
    # main(file_previous, file_latest)
    # paste_picture(comparison_files, dashboard_file)
    # paste_picture(comparison_files, dashboard_file)
